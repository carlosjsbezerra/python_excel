import pandas as pd
import os
from openpyxl import load_workbook

# Caminho para o arquivo Excel dentro da pasta Pasta001
file_path = os.path.join('Pasta001', 'planilha001.xlsx')

# Carregar a planilha Excel especificando o engine openpyxl
df = pd.read_excel(file_path, engine='openpyxl')

# Exibir colunas originais
print("Colunas originais:", df.columns)

# Apagar as colunas especificadas
columns_to_drop = ["UF", "Localização", "Situação SIMEC", "Valor Total do Repasse", "Total Custeio", "Total Capital", "Valor Total", "Data Envio Mec"]
df = df.drop(columns=columns_to_drop)

# Exibir colunas restantes
print("Colunas restantes:", df.columns)

# Manter as colunas desejadas e reordenar se necessário
desired_columns = ["Código INEP", "Nome da Escola", "Município", "Esfera", "Status PDDE"]
df = df[desired_columns]

# Inserir a nova coluna "Regional" entre "Nome da Escola" e "Município"
df.insert(df.columns.get_loc("Município"), "Regional", "")

# Exibir colunas finais
print("Colunas finais:", df.columns)

# Caminho para salvar a nova planilha Excel dentro da pasta Pasta002
new_file_path = os.path.join('Pasta002', 'nova_planilha.xlsx')

# Criar o ExcelWriter com openpyxl
with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
    # Salvar o DataFrame na aba 'Pasta001'
    df.to_excel(writer, index=False, sheet_name='Pasta001')

    # Criar a aba 'referencia_escolas' se não existir
    if 'referencia_escolas' not in writer.book.sheetnames:
        writer.book.create_sheet(title='referencia_escolas')

    # Adicionar colunas 'MUNICIPIO' e 'REGIONAL' na aba 'referencia_escolas'
    ws = writer.book['referencia_escolas']
    ws.append(["MUNICIPIO", "REGIONAL"])
    Pasta001 = [
        ("ANANÁS", "ARAGUAINA"),  #ARAGUAINA
        ("ARAGOMINAS", "ARAGUAINA"),
        ("ARAGUAÍNA", "ARAGUAINA"),
        ("ARAGUANÃ", "ARAGUAINA"),
        ("BABAÇULÂNDIA", "ARAGUAINA"),
        ("BARRA DO OURO", "ARAGUAINA"),
        ("CAMPOS LINDOS", "ARAGUAINA"),
        ("CARMOLÂNDIA", "ARAGUAINA"),
        ("FILADÉLFIA", "ARAGUAINA"),
        ("GOIATINS", "ARAGUAINA"),
        ("MURICILÂNDIA", "ARAGUAINA"),
        ("NOVA OLINDA", "ARAGUAINA"),
        ("NOVO ALEGRE", "ARAGUAINA"),
        ("PIRAQUÊ", "ARAGUAINA"),
        ("RIACHINHO", "ARAGUAINA"),
        ("SANTA FÉ DO ARAGUAIA", "ARAGUAINA"),
        ("WANDERLÂNDIA", "ARAGUAINA"),
        ("XAMBIOÁ", "ARAGUAINA"),
        ("ARAGUATINS", "ARAGUATINS"), #ARAGUATINS
        ("AUGUSTINÓPOLIS", "ARAGUATINS"),
        ("AXIXÁ DO TOCANTINS", "ARAGUATINS"),
        ("BURITI DO TOCANTINS", "ARAGUATINS"),
        ("CARRASCO BONITO", "ARAGUATINS"),
        ("ESPERANTINA", "ARAGUATINS"),
        ("PRAIA NORTE", "ARAGUATINS"),
        ("SAMPAIO", "ARAGUATINS"),
        ("SÃO BENTO DO TOCANTINS", "ARAGUATINS"),
        ("SÃO MIGUEL DO TOCANTINS", "ARAGUATINS"),
        ("SÃO SEBASTIÃO DO TOCANTINS", "ARAGUATINS"),
        ("SÍTIO NOVO DO TOCANTINS", "ARAGUATINS"),
        ("ARRAIAS", "ARRAIAS"), # ARRAIAS
        ("AURORA DO TOCANTINS", "ARRAIAS"),
        ("COMBINADO", "ARRAIAS"),
        ("LAVANDEIRA", "ARRAIAS"),
        ("PARANÃ", "ARRAIAS"),
        ("ARAPOEMA", "COLINAS"), #COLINAS
        ("BANDEIRANTES DO TOCANTINS", "COLINAS"),
        ("BERNARDO SAYÃO", "COLINAS"),
        ("BRASILÂNDIA DO TOCANTINS", "COLINAS"),
        ("COLINAS DO TOCANTINS", "COLINAS"),
        ("ITAPIRATINS", "COLINAS"),
        ("JUARINA", "COLINAS"),
        ("PALMEIRANTE", "COLINAS"),
        ("PAU D'ARCO", "COLINAS"),
        ("TUPIRATINS", "COLINAS"),
        ("ALMAS", "DIANOPOLIS"), #DIANOPOLIS
        ("CONCEIÇÃO DO TOCANTINS", "DIANOPOLIS"),
        ("DIANÓPOLIS", "DIANOPOLIS"),
        ("NOVO JARDIM", "DIANOPOLIS"),
        ("PONTE ALTA DO BOM JESUS", "DIANOPOLIS"),
        ("PORTO ALEGRE DO TOCANTINS", "DIANOPOLIS"),
        ("RIO DA CONCEIÇÃO", "DIANOPOLIS"),
        ("TAGUATINGA", "DIANOPOLIS"),
        ("TAIPAS DO TOCANTINS", "DIANOPOLIS"),
        ("COLMÉIA", "GUARAI"), #GUARAI
        ("COUTO MAGALHÃES", "GUARAI"),
        ("GOIANORTE", "GUARAI"),
        ("GUARAÍ", "GUARAI"),
        ("ITAPORÃ DO TOCANTINS", "GUARAI"),
        ("PEQUIZEIRO", "GUARAI"),
        ("PRESIDENTE KENNEDY", "GUARAI"),
        ("TABOCÃO", "GUARAI"),
        ("ALIANÇA DO TOCANTINS", "GURUPI"),
        ("ALVORADA", "GURUPI"),
        ("ARAGUAÇU", "GURUPI"),
        ("CARIRI DO TOCANTINS", "GURUPI"),
        ("CRIXÁS DO TOCANTINS", "GURUPI"),
        ("DUERÉ", "GURUPI"),
        ("FIGUEIRÓPOLIS", "GURUPI"),
        ("FORMOSO DO ARAGUAIA", "GURUPI"),
        ("GURUPI", "GURUPI"),
        ("JAÚ DO TOCANTINS", "GURUPI"),
        ("PALMEIRÓPOLIS", "GURUPI"),
        ("PEIXE", "GURUPI"),
        ("SANDOLÂNDIA", "GURUPI"),
        ("SÃO SALVADOR DO TOCANTINS", "GURUPI"),
        ("SÃO VALÉRIO", "GURUPI"),
        ("SUCUPIRA", "GURUPI"),
        ("TALISMÃ", "GURUPI"),
        ("DOIS IRMÃOS DO TOCANTINS", "MIRACEMA"),
        ("LIZARDA", "MIRACEMA"),
        ("MIRACEMA DO TOCANTINS", "MIRACEMA"),
        ("MIRANORTE", "MIRACEMA"),
        ("RIO DOS BOIS", "MIRACEMA"),
        ("TOCANTÍNIA", "MIRACEMA"),
        ("APARECIDA DO RIO NEGRO", "PALMAS"),
        ("LAGOA DO TOCANTINS", "PALMAS"),
        ("LAJEADO", "PALMAS"),
        ("MATEIROS", "PALMAS"),
        ("NOVO ACORDO", "PALMAS"),
        ("PALMAS", "PALMAS"),
        ("RIO SONO", "PALMAS"),
        ("SANTA TEREZA DO TOCANTINS", "PALMAS"),
        ("SÃO FÉLIX DO TOCANTINS", "PALMAS"),
        ("ABREULÂNDIA", "PARAISO"),
        ("ARAGUACEMA", "PARAISO"),
        ("BARROLÂNDIA", "PARAISO"),
        ("CASEARA", "PARAISO"),
        ("CRISTALÂNDIA", "PARAISO"),
        ("DIVINÓPOLIS DO TOCANTINS", "PARAISO"),
        ("LAGOA DA CONFUSÃO", "PARAISO"),
        ("MARIANÓPOLIS DO TOCANTINS", "PARAISO"),
        ("NOVA ROSALÂNDIA", "PARAISO"),
        ("PARAÍSO DO TOCANTINS", "PARAISO"),
        ("PIUM", "PARAISO"),
        ("PUGMIL", "PARAISO"),
        ("BOM JESUS DO TOCANTINS", "PEDRO AFONSO"),
        ("CENTENÁRIO", "PEDRO AFONSO"),
        ("ITACAJÁ", "PEDRO AFONSO"),
        ("PEDRO AFONSO", "PEDRO AFONSO"),
        ("RECURSOLÂNDIA", "PEDRO AFONSO"),
        ("SANTA MARIA DO TOCANTINS", "PEDRO AFONSO"),
        ("TUPIRAMA", "PEDRO AFONSO"),
        ("BREJINHO DE NAZARÉ", "PORTO NACIONAL"),
        ("CHAPADA DA NATIVIDADE", "PORTO NACIONAL"),
        ("FÁTIMA", "PORTO NACIONAL"),
        ("IPUEIRAS", "PORTO NACIONAL"),
        ("MONTE DO CARMO", "PORTO NACIONAL"),
        ("NATIVIDADE", "PORTO NACIONAL"),
        ("OLIVEIRA DE FÁTIMA", "PORTO NACIONAL"),
        ("PINDORAMA DO TOCANTINS", "PORTO NACIONAL"),
        ("PONTE ALTA DO TOCANTINS", "PORTO NACIONAL"),
        ("PORTO NACIONAL", "PORTO NACIONAL"),
        ("SANTA RITA DO TOCANTINS", "PORTO NACIONAL"),
        ("SANTA ROSA DO TOCANTINS", "PORTO NACIONAL"),
        ("SILVANÓPOLIS", "PORTO NACIONAL"),
        ("AGUIARNÓPOLIS", "TOCANTINOPOLIS"),
        ("ANGICO", "TOCANTINOPOLIS"),
        ("CACHOEIRINHA", "TOCANTINOPOLIS"),
        ("DARCINÓPOLIS", "TOCANTINOPOLIS"),
        ("ITAGUATINS", "TOCANTINOPOLIS"),
        ("LUZINÓPOLIS", "TOCANTINOPOLIS"),
        ("MAURILÂNDIA DO TOCANTINS", "TOCANTINOPOLIS"),
        ("NAZARÉ", "TOCANTINOPOLIS"),
        ("PALMEIRAS DO TOCANTINS", "TOCANTINOPOLIS"),
        ("SANTA TEREZINHA DO TOCANTINS", "TOCANTINOPOLIS"),
        ("TOCANTINÓPOLIS", "TOCANTINOPOLIS")
    ]
    for dado in Pasta001:
        ws.append(dado)
    
    # Adicionar a função VLOOKUP nas células C2 até C498 da planilha 'Pasta001'
    ws = writer.book['Pasta001']
    for row in range(2, 499):
        ws[f'C{row}'] = f'=VLOOKUP(D{row}, referencia_escolas!$A$1:$B$138, 2, 0)'
    

    # Adicionar filtros às colunas especificadas
    ws.auto_filter.ref = ws.dimensions



print(f"Planilha salva em {new_file_path} com as abas 'Pasta001' e 'referencia_escolas'")
