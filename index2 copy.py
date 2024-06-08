from openpyxl import load_workbook

# Carrega a planilha "nova_planilha"
wb = load_workbook(filename='Pasta002/nova_planilha.xlsx')

# Obtém a subplanilha "Escolas" e a planilha "planilha002" e a subplanilha "SRE ARRAIAS"
ws_escolas = wb['Escolas']
wb_planilha002 = load_workbook(filename='Pasta002/planilha002.xlsx')
ws_sre_arraias = wb_planilha002['SRE ARRAIAS']

# Copia os valores da subplanilha "Escolas" para a subplanilha "SRE ARRAIAS"
for row in ws_escolas.iter_rows(values_only=True):
    ws_sre_arraias.append(row)

# Salva as alterações na planilha "planilha002"
wb_planilha002.save(filename='Pasta002/planilha002.xlsx')
