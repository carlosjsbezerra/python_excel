from openpyxl import load_workbook

def organizar_coluna_f():
    # Mapeamento dos status para a ordem desejada
    status_ordem = {
        'Não iniciado': 1,
        'Em cadastramento': 2,
        'Em análise do MEC': 3
    }

    # Carregar a planilha "planilha002" e obter a subplanilha "SRE ARRAIAS"
    wb_planilha002 = load_workbook(filename='Pasta002/planilha002.xlsx')
    ws_sre_pedro_arraias = wb_planilha002['SRE ARRAIAS']

    # Inicializa uma lista para armazenar os valores da coluna "F"
    valores_coluna_f = []

    # Percorrer as linhas da coluna "F" da linha 11 até a linha 26
    for row in ws_sre_pedro_arraias.iter_rows(min_row=11, max_row=26, min_col=6, max_col=6, values_only=True):
        valores_coluna_f.append(row[0])

    # Ordenar os valores de acordo com a ordem especificada
    valores_coluna_f.sort(key=lambda x: status_ordem.get(x, 0))

    # Atualizar os valores na coluna "F" mantendo a ordem
    for i, valor in enumerate(valores_coluna_f):
        ws_sre_pedro_arraias.cell(row=11 + i, column=6, value=valor)

    # Salvar as alterações na planilha "planilha002"
    wb_planilha002.save(filename='Pasta002/planilha002.xlsx')

    print("Coluna 'F' organizada com sucesso na subplanilha 'SRE ARRAIAS' da linha 11 até a linha 26.")
