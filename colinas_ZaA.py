from openpyxl import load_workbook

def organizar_coluna_f():

    start_line = 11
    finish_line = 29

    # Mapeamento dos status para a ordem desejada
    status_ordem = {
        'Não iniciado': 1,
        'Em cadastramento': 2,
        'Em análise do MEC': 3
    }

    # Carregar a planilha "planilha002" e obter a subplanilha "SRE COLINAS DO TOCANTINS"
    wb_planilha002 = load_workbook(filename='PIEC_2024_v2/lista_escola_selecionada.xlsx')
    ws_sre_colinas = wb_planilha002['SRE COLINAS DO TOCANTINS']

    # Inicializa uma lista para armazenar os valores da coluna "F"
    valores_coluna_f = []

    # Percorrer as linhas da coluna "F" da linha 11 até a linha 26
    for row in ws_sre_colinas.iter_rows(min_row=start_line, max_row=finish_line, min_col=6, max_col=6, values_only=True):
        valores_coluna_f.append(row[0])

    # Ordenar os valores de acordo com a ordem especificada
    valores_coluna_f.sort(key=lambda x: status_ordem.get(x, 0))

    # Atualizar os valores na coluna "F" mantendo a ordem
    for i, valor in enumerate(valores_coluna_f):
        ws_sre_colinas.cell(row=11 + i, column=6, value=valor)

    # Salvar as alterações na planilha "planilha002"
    wb_planilha002.save(filename='PIEC_2024_v2/lista_escola_selecionada.xlsx')

    print("Coluna 'F' organizada com sucesso na subplanilha 'SRE COLINAS DO TOCANTINS' da linha 11 até a linha 29.")
