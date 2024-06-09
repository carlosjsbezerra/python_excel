from openpyxl import load_workbook

from paraiso_ZaA import organizar_coluna_f

def copiar_linhas_para_sre(sre_nome, valor_coluna_c_procurado):
    # Carregar a planilha "nova_planilha"
    wb_escolas = load_workbook(filename='Pasta002/nova_planilha.xlsx', data_only=True)
    print("Subplanilhas em nova_planilha.xlsx:", wb_escolas.sheetnames)

    # Obter a subplanilha "Escolas"
    ws_escolas = wb_escolas['Escolas']

    # Carregar a planilha "planilha002" e obter a subplanilha especificada
    wb_planilha002 = load_workbook(filename='Pasta002/planilha002.xlsx')
    print("Subplanilhas em planilha002.xlsx:", wb_planilha002.sheetnames)

    ws_sre = wb_planilha002[sre_nome]

    # Inicializa uma lista para armazenar as linhas que serão copiadas
    linhas_para_copiar = []

    # Iterar sobre as linhas da subplanilha "Escolas"
    for row in ws_escolas.iter_rows(min_row=2, values_only=True):
        valor_coluna_c = row[2]  # Valor na coluna C
        # Imprimir o valor da coluna C para cada linha
        print("Valor na coluna C:", valor_coluna_c)
        if valor_coluna_c == valor_coluna_c_procurado:  # Verificar se o valor na coluna "C" é o procurado
            linhas_para_copiar.append(row)
            print("Linha encontrada:", row)

    # Verificar se alguma linha foi encontrada
    if not linhas_para_copiar:
        print(f"Nenhuma linha encontrada com o valor '{valor_coluna_c_procurado}' na coluna C.")

    # Adicionar as linhas filtradas começando na célula A11 da subplanilha especificada
    start_row = 11
    for i, linha in enumerate(linhas_para_copiar):
        for j, value in enumerate(linha):
            ws_sre.cell(row=start_row + i, column=j + 1, value=value)

    # Verificar se as linhas foram adicionadas
    if linhas_para_copiar:
        print(f"{len(linhas_para_copiar)} linhas copiadas para '{sre_nome}'.")

    # Salvar as alterações na planilha "planilha002"
    wb_planilha002.save(filename='Pasta002/planilha002.xlsx')

    print(f"Linhas copiadas com sucesso para a subplanilha '{sre_nome}'.")

# Chamada da função
copiar_linhas_para_sre('SRE PARAISO', 'PARAISO')
# Chamar a função organizar_coluna_f
organizar_coluna_f()