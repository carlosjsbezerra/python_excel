from openpyxl import load_workbook

from araguaina_ZaA import organizar_coluna_f

def copiar_linhas_araguaina():
    # Carregar a planilha "nova_planilha"
    wb_escolas = load_workbook(filename='./nova_planilha.xlsx', data_only=True)

    # Obter a subplanilha "Escolas"
    ws_escolas = wb_escolas['Escolas']

    # Códigos INEP para os quais você não quer copiar
    codigos_inep = [
        "17028507", "17051452", "17029341",
        "17044049", "17028493", "17056276"
    ]

    # Inicializa uma lista para armazenar as linhas que serão copiadas
    linhas_para_copiar = []

    # Iterar sobre as linhas da subplanilha "Escolas"
    for row in ws_escolas.iter_rows(min_row=2, values_only=True):
        codigo_inep = str(row[0])  # Valor na coluna Código INEP
        regional = row[2]  # Valor na coluna Regional
        if regional == 'ARAGUAINA' and codigo_inep not in codigos_inep:
            linhas_para_copiar.append(row)

    # Carregar a planilha "planilha002" e obter a subplanilha "SRE ARAGUAINA"
    wb_planilha002 = load_workbook(filename='PIEC_2024_v2/lista_escola_selecionada.xlsx')
    ws_sre_araguaina = wb_planilha002['SRE ARAGUAINA']

    # Adicionar as linhas filtradas começando na célula A11 da subplanilha "SRE ARAGUAINA"
    start_row = 11
    for i, linha in enumerate(linhas_para_copiar):
        for j, value in enumerate(linha):
            ws_sre_araguaina.cell(row=start_row + i, column=j + 1, value=value)

    # Salvar as alterações na planilha "planilha002"
    wb_planilha002.save(filename='PIEC_2024_v2/lista_escola_selecionada.xlsx')

    print("Linhas copiadas com sucesso para a subplanilha 'SRE ARAGUAINA'.")

    # Chamar a função organizar_coluna_f
    organizar_coluna_f()


#copiar_linhas_araguaina()
# Chamar a função para executar o processo
#organizar_coluna_f()
