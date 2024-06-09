from openpyxl import load_workbook

# Códigos INEP para os quais você quer copiar
codigos_inep = [
    "17028507", "17051452", "17029341",
    "17044049", "17028493", "17056276"
]

# Carregar a planilha "nova_planilha"
wb_nova_planilha = load_workbook(filename='Pasta002/nova_planilha.xlsx', data_only=True)
print("Subplanilhas em nova_planilha.xlsx:", wb_nova_planilha.sheetnames)

# Obter a subplanilha "Escolas" na planilha "nova_planilha"
ws_escolas = wb_nova_planilha['Escolas']

# Carregar a planilha "planilha002" e obter a subplanilha "SRE PEDRO AFONSO"
wb_planilha002 = load_workbook(filename='Pasta002/planilha002.xlsx')
print("Subplanilhas em planilha002.xlsx:", wb_planilha002.sheetnames)

ws_sre_pedro_afonso = wb_planilha002['SRE PEDRO AFONSO']

# Inicializa uma lista para armazenar as linhas que serão copiadas
linhas_para_copiar = []

# Iterar sobre as linhas da subplanilha "Escolas"
for row in ws_escolas.iter_rows(min_row=2, values_only=True):
    valor_coluna_c = row[2]  # Valor na coluna C
    # Imprimir o valor da coluna C para cada linha
    print("Valor na coluna C:", valor_coluna_c)
    if valor_coluna_c == 'PEDRO AFONSO':  # Verificar se o valor na coluna "C" é "PEDRO AFONSO"
        linhas_para_copiar.append(row)
        print("Linha encontrada:", row)

# Verificar se alguma linha foi encontrada
if not linhas_para_copiar:
    print("Nenhuma linha encontrada com o valor 'PEDRO AFONSO' na coluna C.")

# Adicionar as linhas filtradas começando na célula A11 da subplanilha "SRE PEDRO AFONSO"
start_row = 11
for i, linha in enumerate(linhas_para_copiar):
    for j, value in enumerate(linha):
        ws_sre_pedro_afonso.cell(row=start_row + i, column=j + 1, value=value)

# Adicionar as linhas filtradas à subplanilha "SRE PEDRO AFONSO" na coluna "A" a partir da linha 35
for row in ws_escolas.iter_rows(min_row=2, values_only=True):
    codigo_inep = str(row[0])  # Valor na coluna Código INEP
    if codigo_inep in codigos_inep:
        linhas_para_copiar.append(row)

# Adicionar as linhas filtradas à subplanilha "SRE PEDRO AFONSO" na coluna "A" a partir da linha 35
start_row = 11
for i, linha in enumerate(linhas_para_copiar):
    for j, value in enumerate(linha):
        ws_sre_pedro_afonso.cell(row=start_row + i, column=j + 1, value=value)

# Salvar as alterações na planilha "planilha002"
wb_planilha002.save(filename='Pasta002/planilha002.xlsx')

print("Linhas copiadas com sucesso para a subplanilha 'SRE PEDRO AFONSO'.")
