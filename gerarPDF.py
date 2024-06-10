import os
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

def gerar_pdfs_subplanilhas(arquivo_excel, diretorio_saida):
    # Carregar a planilha do Excel
    wb = load_workbook(arquivo_excel)

    # Verificar se o diretório de saída existe, caso contrário, criar
    if not os.path.exists(diretorio_saida):
        os.makedirs(diretorio_saida)

    # Iterar sobre todas as subplanilhas
    for sheetname in wb.sheetnames:
        print(f"Gerando PDF da subplanilha: {sheetname}")
        sheet = wb[sheetname]
        caminho_pdf = os.path.join(diretorio_saida, f"{sheetname}.pdf")

        # Criar um canvas para o PDF
        c = canvas.Canvas(caminho_pdf, pagesize=letter)
        largura, altura = letter
        linha_altura = altura - 40  # Margem superior

        # Iterar sobre as linhas da subplanilha
        for row in sheet.iter_rows(values_only=True):
            linha_texto = "  ".join([str(cell) for cell in row if cell is not None])
            c.drawString(30, linha_altura, linha_texto)
            linha_altura -= 15  # Altura da linha para a próxima

            # Adicionar nova página se ultrapassar o tamanho da página
            if linha_altura < 40:
                c.showPage()
                linha_altura = altura - 40

        c.save()
        print(f"PDF salvo em: {caminho_pdf}")

# Nome do arquivo Excel com as subplanilhas
arquivo_excel = './lista_escola_selecionada.xlsx'

# Diretório onde os PDFs serão salvos
diretorio_saida = './PDF'

# Chamada da função para gerar os PDFs
gerar_pdfs_subplanilhas(arquivo_excel, diretorio_saida)
